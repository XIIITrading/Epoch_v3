"""
Epoch Backtest Journal - Excel Export Utility
Manages Excel workbook with multiple analysis worksheets.
"""

import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Union
import pandas as pd
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from config.settings import REPORTS_DIR
from config.theme import COLORS


class ExcelExporter:
    """
    Manages Excel workbook for backtest journal analysis.
    Each module adds its own worksheet with raw data and results.
    """

    # Style definitions matching dark theme
    STYLES = {
        'header': {
            'font': Font(bold=True, color='FFFFFF', size=11),
            'fill': PatternFill(start_color='2A2A4E', end_color='2A2A4E', fill_type='solid'),
            'border': Border(
                bottom=Side(style='thin', color='333333'),
                right=Side(style='thin', color='333333')
            ),
            'alignment': Alignment(horizontal='center', vertical='center')
        },
        'subheader': {
            'font': Font(bold=True, color='E0E0E0', size=10),
            'fill': PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center')
        },
        'data': {
            'font': Font(color='E0E0E0', size=10),
            'fill': PatternFill(start_color='0F0F1A', end_color='0F0F1A', fill_type='solid'),
            'border': Border(
                bottom=Side(style='thin', color='222222'),
                right=Side(style='thin', color='222222')
            ),
            'alignment': Alignment(horizontal='center', vertical='center')
        },
        'data_alt': {
            'font': Font(color='E0E0E0', size=10),
            'fill': PatternFill(start_color='12121F', end_color='12121F', fill_type='solid'),
            'border': Border(
                bottom=Side(style='thin', color='222222'),
                right=Side(style='thin', color='222222')
            ),
            'alignment': Alignment(horizontal='center', vertical='center')
        },
        'positive': {
            'font': Font(color='26A69A', size=10, bold=True),
        },
        'negative': {
            'font': Font(color='EF5350', size=10, bold=True),
        },
        'title': {
            'font': Font(bold=True, color='FFFFFF', size=14),
            'fill': PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        },
        'total_row': {
            'font': Font(bold=True, color='FFFFFF', size=10),
            'fill': PatternFill(start_color='3A3A5E', end_color='3A3A5E', fill_type='solid'),
            'border': Border(
                top=Side(style='medium', color='444444'),
                bottom=Side(style='medium', color='444444')
            ),
        }
    }

    def __init__(self, workbook_path: Path = None):
        """
        Initialize Excel exporter.

        Args:
            workbook_path: Path to workbook. Auto-generates if not provided.
        """
        self.workbook_path = workbook_path
        self.workbook = None
        self._is_new = False

    def open(self, create_if_missing: bool = True) -> 'ExcelExporter':
        """
        Open or create workbook.

        Args:
            create_if_missing: Create new workbook if path doesn't exist

        Returns:
            Self for chaining
        """
        if self.workbook_path and self.workbook_path.exists():
            self.workbook = openpyxl.load_workbook(self.workbook_path)
            self._is_new = False
        elif create_if_missing:
            self.workbook = openpyxl.Workbook()
            self._is_new = True
            # Remove default sheet if we'll add our own
            if 'Sheet' in self.workbook.sheetnames:
                del self.workbook['Sheet']
        else:
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

        return self

    def save(self, path: Path = None) -> Path:
        """
        Save workbook.

        Args:
            path: Save path. Uses self.workbook_path if not provided.

        Returns:
            Path where workbook was saved
        """
        save_path = path or self.workbook_path
        if save_path is None:
            raise ValueError("No save path specified")

        self.workbook.save(save_path)
        return save_path

    def close(self):
        """Close workbook without saving."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def add_worksheet(
        self,
        name: str,
        title: str = None,
        overwrite: bool = True
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """
        Add or get a worksheet.

        Args:
            name: Worksheet name (max 31 chars)
            title: Optional title to display at top of sheet
            overwrite: If True, clear existing sheet. If False, return existing.

        Returns:
            Worksheet object
        """
        # Truncate name to Excel limit
        name = name[:31]

        if name in self.workbook.sheetnames:
            if overwrite:
                del self.workbook[name]
                ws = self.workbook.create_sheet(name)
            else:
                ws = self.workbook[name]
        else:
            ws = self.workbook.create_sheet(name)

        # Set dark background color for entire sheet
        ws.sheet_properties.tabColor = "1A1A2E"

        # Add title if provided
        if title:
            ws.merge_cells('A1:H1')
            cell = ws['A1']
            cell.value = title
            self._apply_style(cell, 'title')
            ws.row_dimensions[1].height = 30

        return ws

    def write_dataframe(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        df: pd.DataFrame,
        start_row: int = 1,
        start_col: int = 1,
        include_header: bool = True,
        include_index: bool = False,
        header_style: str = 'header',
        highlight_columns: Dict[str, str] = None
    ) -> int:
        """
        Write DataFrame to worksheet with styling.

        Args:
            ws: Worksheet to write to
            df: DataFrame to write
            start_row: Starting row (1-indexed)
            start_col: Starting column (1-indexed)
            include_header: Include column headers
            include_index: Include DataFrame index
            header_style: Style name for headers
            highlight_columns: Dict mapping column names to 'positive'/'negative' for conditional coloring

        Returns:
            Next available row after data
        """
        highlight_columns = highlight_columns or {}

        rows = list(dataframe_to_rows(df, index=include_index, header=include_header))

        current_row = start_row
        for i, row in enumerate(rows):
            # Skip empty row that dataframe_to_rows sometimes adds
            if row == [None] * len(row):
                continue

            is_header = (i == 0 and include_header)

            for j, value in enumerate(row):
                cell = ws.cell(row=current_row, column=start_col + j, value=value)

                if is_header:
                    self._apply_style(cell, header_style)
                else:
                    # Alternate row colors
                    style = 'data' if (current_row - start_row) % 2 == 0 else 'data_alt'
                    self._apply_style(cell, style)

                    # Apply conditional coloring for specific columns
                    if include_header and i > 0:
                        col_name = rows[0][j] if j < len(rows[0]) else None
                        if col_name in highlight_columns:
                            if isinstance(value, (int, float)):
                                if value > 0:
                                    cell.font = self.STYLES['positive']['font']
                                elif value < 0:
                                    cell.font = self.STYLES['negative']['font']

            current_row += 1

        # Auto-size columns
        self._auto_size_columns(ws, start_col, start_col + len(df.columns) - 1)

        return current_row

    def write_summary_table(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        data: List[Dict[str, Any]],
        columns: List[str],
        start_row: int = 1,
        start_col: int = 1,
        title: str = None,
        total_row: Dict[str, Any] = None,
        highlight_columns: List[str] = None
    ) -> int:
        """
        Write a summary table with optional title and total row.

        Args:
            ws: Worksheet to write to
            data: List of dicts with row data
            columns: Column names/keys to include
            start_row: Starting row
            start_col: Starting column
            title: Optional section title
            total_row: Optional dict with total row data
            highlight_columns: Columns to apply positive/negative coloring

        Returns:
            Next available row after table
        """
        highlight_columns = highlight_columns or []
        current_row = start_row

        # Write title
        if title:
            ws.merge_cells(
                start_row=current_row,
                start_column=start_col,
                end_row=current_row,
                end_column=start_col + len(columns) - 1
            )
            cell = ws.cell(row=current_row, column=start_col, value=title)
            self._apply_style(cell, 'subheader')
            current_row += 1

        # Write headers
        for j, col in enumerate(columns):
            cell = ws.cell(row=current_row, column=start_col + j, value=col)
            self._apply_style(cell, 'header')
        current_row += 1

        # Write data rows
        for i, row_data in enumerate(data):
            style = 'data' if i % 2 == 0 else 'data_alt'
            for j, col in enumerate(columns):
                value = row_data.get(col, '')
                cell = ws.cell(row=current_row, column=start_col + j, value=value)
                self._apply_style(cell, style)

                # Highlight if needed
                if col in highlight_columns and isinstance(value, (int, float)):
                    if value > 0:
                        cell.font = self.STYLES['positive']['font']
                    elif value < 0:
                        cell.font = self.STYLES['negative']['font']

            current_row += 1

        # Write total row
        if total_row:
            for j, col in enumerate(columns):
                value = total_row.get(col, '')
                cell = ws.cell(row=current_row, column=start_col + j, value=value)
                self._apply_style(cell, 'total_row')

                if col in highlight_columns and isinstance(value, (int, float)):
                    if value > 0:
                        cell.font = Font(bold=True, color='26A69A', size=10)
                    elif value < 0:
                        cell.font = Font(bold=True, color='EF5350', size=10)

            current_row += 1

        # Auto-size columns
        self._auto_size_columns(ws, start_col, start_col + len(columns) - 1)

        return current_row + 1  # Add blank row after table

    def write_raw_trades(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        trades: List[Dict],
        start_row: int = 1
    ) -> int:
        """
        Write raw trade data to worksheet.

        Args:
            ws: Worksheet
            trades: List of trade dicts
            start_row: Starting row

        Returns:
            Next available row
        """
        if not trades:
            return start_row

        # Convert to DataFrame
        df = pd.DataFrame(trades)

        # Select and order columns
        columns = [
            'trade_id', 'date', 'ticker', 'model', 'zone_type', 'direction',
            'entry_price', 'entry_time', 'exit_price', 'exit_time',
            'exit_reason', 'pnl_r', 'is_winner'
        ]

        # Only include columns that exist
        columns = [c for c in columns if c in df.columns]
        df = df[columns]

        return self.write_dataframe(
            ws, df,
            start_row=start_row,
            highlight_columns={'pnl_r': 'value'}
        )

    def _apply_style(self, cell, style_name: str):
        """Apply a named style to a cell."""
        style = self.STYLES.get(style_name, {})
        if 'font' in style:
            cell.font = style['font']
        if 'fill' in style:
            cell.fill = style['fill']
        if 'border' in style:
            cell.border = style['border']
        if 'alignment' in style:
            cell.alignment = style['alignment']

    def _auto_size_columns(self, ws, start_col: int, end_col: int):
        """Auto-size columns based on content."""
        for col_idx in range(start_col, end_col + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width


def generate_workbook_path(
    start_date: date = None,
    end_date: date = None,
    prefix: str = "bt_journal"
) -> Path:
    """
    Generate a workbook path based on date range.

    Args:
        start_date: Analysis start date
        end_date: Analysis end date
        prefix: Filename prefix

    Returns:
        Path for workbook
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if start_date and end_date:
        if start_date == end_date:
            date_str = start_date.strftime("%Y-%m-%d")
        else:
            date_str = f"{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}"
    else:
        date_str = "all_data"

    filename = f"{prefix}_{date_str}_{timestamp}.xlsx"
    return REPORTS_DIR / filename
